from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _bool_from_str(value: object) -> bool:
    if value is None:
        return False
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "t"}:
        return True
    if normalized in {"false", "0", "no", "n", "f"}:
        return False
    return False


def build_ssl_dataset(
    master_df: pd.DataFrame, raw_df: pd.DataFrame, bu: str, ssl: str
) -> pd.DataFrame:
    """
    Returns a row-per-employee dataframe for the given BU/SSL.
    Base rows come from master roster filtered to active=True.
    Join weekly raw_df by GPN.
    Computes status and output columns.
    """
    master = master_df.copy()
    master["gpn"] = master["gpn"].astype(str).str.strip()
    master["active_bool"] = master["active"].apply(_bool_from_str)
    master["on_leave_bool"] = master["on_leave"].apply(_bool_from_str)

    active_master = master[master["active_bool"]].copy()
    active_master = active_master[
        (active_master["bu"].astype(str).str.strip() == str(bu))
        & (active_master["ssl"].astype(str).str.strip() == str(ssl))
    ].copy()

    raw = raw_df.copy()
    raw["GPN"] = raw["GPN"].astype(str).str.strip()

    # Keep only required raw columns to avoid merge collisions
    raw_cols = [
        "GPN",
        "Competency",
        "Employee Status",
        "Effective Available Hours",
        "Chargeable Hours",
        "util",
        "missing_timesheet",
        "vacation",
        "inactive_leave",
        "unmapped_competency",
        "Rank Description",
    ]
    raw_cols = [c for c in raw_cols if c in raw.columns]
    raw_small = raw[raw_cols].copy()

    merged = active_master.merge(
        raw_small,
        left_on="gpn",
        right_on="GPN",
        how="left",
    )

    missing_in_export = merged["GPN"].isna()

    on_leave = (
        merged["on_leave_bool"]
        | merged.get("inactive_leave", False).fillna(False)
        | missing_in_export
    )
    vacation = merged.get("vacation", False).fillna(False)
    missing_ts = merged.get("missing_timesheet", False).fillna(False)
    unmapped = merged.get("unmapped_competency", False).fillna(False)

    status = pd.Series(["Normal"] * len(merged))
    status = status.mask(unmapped, "Unmapped")
    status = status.mask(missing_ts, "Missing timesheet")
    status = status.mask(vacation, "Vacation")
    status = status.mask(on_leave, "On leave")

    detail = pd.DataFrame({
        "gpn": merged["gpn"],
        "display_name": merged.get("display_name", ""),
        "rank_bucket": merged.get("rank_bucket", ""),
        "bu": merged.get("bu", ""),
        "ssl": merged.get("ssl", ""),
        "competency": merged.get("Competency", ""),
        "employee_status": merged.get("Employee Status", ""),
        "effective_available_hours": merged.get("Effective Available Hours", pd.NA),
        "chargeable_hours": merged.get("Chargeable Hours", pd.NA),
        "util": merged.get("util", pd.NA),
        "missing_timesheet": missing_ts.astype(bool),
        "vacation": vacation.astype(bool),
        "inactive_leave": merged.get("inactive_leave", False).fillna(False).astype(bool),
        "unmapped_competency": unmapped.astype(bool),
        "on_leave_master": merged["on_leave_bool"].astype(bool),
        "status": status,
        "notes": merged.get("notes", ""),
    })

    return detail


def build_ssl_summary(
    totals_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    bu: str,
    ssl: str,
    week_meta: Dict,
) -> pd.DataFrame:
    """
    Returns a small summary dataframe (key/value rows).
    """
    totals = totals_df.copy()
    totals = totals[
        (totals["BU"].astype(str).str.strip() == str(bu))
        & (totals["SSL"].astype(str).str.strip() == str(ssl))
    ]
    totals_row = totals.iloc[0] if len(totals) > 0 else pd.Series({})

    raw = raw_df.copy()
    if "GPN" in raw.columns:
        raw_gpns = set(raw["GPN"].dropna().astype(str).str.strip())
    else:
        raw_gpns = set()

    master_df = week_meta.get("master_df")
    if master_df is None:
        raise ValueError("master_df missing from week_meta for summary computation")

    detail = build_ssl_dataset(master_df=master_df, raw_df=raw_df, bu=bu, ssl=ssl)
    headcount = len(detail)
    present = detail["gpn"].astype(str).str.strip().isin(raw_gpns).sum()
    missing = headcount - present
    missing_ts = detail["missing_timesheet"].sum()
    vacation = detail["vacation"].sum()
    on_leave = (detail["status"] == "On leave").sum()

    summary_rows = [
        ("Week key", week_meta.get("short_key", "")),
        ("Export format", week_meta.get("export_format", "")),
        ("Period", f"{week_meta.get('start_date', '')} - {week_meta.get('end_date', '')}"),
        ("Util all", totals_row.get("util_all", pd.NA)),
        ("Util excl missing", totals_row.get("util_excl_missing", pd.NA)),
        ("Headcount (master active)", headcount),
        ("Present in export", int(present)),
        ("Missing from export", int(missing)),
        ("Missing timesheets", int(missing_ts)),
        ("Vacation", int(vacation)),
        ("On leave", int(on_leave)),
    ]

    return pd.DataFrame(summary_rows, columns=["Metric", "Value"])


def write_ssl_excel(detail_df: pd.DataFrame, summary_df: pd.DataFrame, out_path: Path) -> None:
    """
    Writes an .xlsx with at least two sheets:
    - 'Summary'
    - 'Employees'
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        detail_df.to_excel(writer, sheet_name="Employees", index=False)

        wb = writer.book
        ws_summary = writer.sheets["Summary"]
        ws_employees = writer.sheets["Employees"]

        ws_summary.freeze_panes = "A2"

        ws_employees.freeze_panes = "A2"

        # Percent format for utilization in Employees sheet
        if "util" in detail_df.columns:
            util_col = detail_df.columns.get_loc("util") + 1
            for row in range(2, len(detail_df) + 2):
                ws_employees.cell(row=row, column=util_col).number_format = "0.0%"

        # Percent format for summary utilization rows
        for row_idx, metric in enumerate(summary_df["Metric"].tolist(), start=2):
            if metric in {"Util all", "Util excl missing"}:
                ws_summary.cell(row=row_idx, column=2).number_format = "0.0%"

        # -----------------------------
        # Styling: Headers
        # -----------------------------
        header_font = Font(bold=True)
        for cell in ws_summary[1]:
            cell.font = header_font
        for cell in ws_employees[1]:
            cell.font = header_font

        # -----------------------------
        # Make Employees sheet a formatted Excel Table
        # -----------------------------
        last_row = ws_employees.max_row
        last_col = ws_employees.max_column
        table_ref = f"A1:{get_column_letter(last_col)}{last_row}"

        table_name = f"Employees_{out_path.stem}".replace("-", "_").replace(" ", "_")
        table_name = "".join(ch for ch in table_name if ch.isalnum() or ch == "_")
        if len(table_name) > 250:
            table_name = table_name[:250]

        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws_employees.add_table(tab)

        # -----------------------------
        # Number formats for hours columns
        # -----------------------------
        hour_cols = ["effective_available_hours", "chargeable_hours"]
        for col_name in hour_cols:
            if col_name in detail_df.columns:
                col_idx = detail_df.columns.get_loc(col_name) + 1
                for r in range(2, len(detail_df) + 2):
                    ws_employees.cell(row=r, column=col_idx).number_format = "0.0"

        # Wrap long text in notes column
        if "notes" in detail_df.columns:
            notes_col = detail_df.columns.get_loc("notes") + 1
            wrap = Alignment(wrap_text=True, vertical="top")
            for r in range(2, len(detail_df) + 2):
                ws_employees.cell(row=r, column=notes_col).alignment = wrap

        # -----------------------------
        # Auto column widths (with caps)
        # -----------------------------
        def _best_width(values, min_w=10, max_w=45):
            if not values:
                return min_w
            m = max(len(v) for v in values)
            w = m + 2
            return max(min_w, min(max_w, w))

        for col_idx in range(1, last_col + 1):
            col_letter = get_column_letter(col_idx)
            header = ws_employees.cell(row=1, column=col_idx).value
            header = "" if header is None else str(header)

            # sample up to first 200 rows for performance
            sample_vals = [header]
            for r in range(2, min(last_row, 200) + 1):
                v = ws_employees.cell(row=r, column=col_idx).value
                sample_vals.append("" if v is None else str(v))

            max_w = 60 if header in {"display_name", "notes"} else 45
            min_w = 8 if header in {"bu", "ssl"} else 10
            if header in {"missing_timesheet", "vacation", "inactive_leave", "unmapped_competency", "on_leave_master"}:
                max_w = 20
                min_w = 12

            ws_employees.column_dimensions[col_letter].width = _best_width(sample_vals, min_w=min_w, max_w=max_w)

        # Summary sheet widths for readability
        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 40
